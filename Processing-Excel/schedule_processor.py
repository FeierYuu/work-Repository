"""
课程表处理程序
从Excel文件读取教师课程信息，并生成PDF格式的教师时间表
"""

import re
from dataclasses import dataclass
from typing import Dict, List, Optional
import openpyxl
from datetime import datetime


@dataclass
class ScheduleEntry:
    """课程条目"""
    day: str
    time: str
    subject: str
    teacher: str
    room: str
    group: str
    activity_type: str  # лк (лекция), пз (практическое), лб (лабораторная)


class ExcelReader:
    """Excel文件读取器"""

    DAYS_MAP = {
        'ПОНЕДЕЛЬНИК': 'Понедельник',
        'ВТОРНИК': 'Вторник',
        'СРЕДА': 'Среда',
        'ЧЕТВЕРГ': 'Четверг',
        'ПЯТНИЦА': 'Пятница',
        'СУББОТА': 'Суббота'
    }

    def __init__(self, file_path: str):
        self.file_path = file_path
        self.workbook = None
        self.worksheet = None

    def load(self):
        """加载Excel文件"""
        self.workbook = openpyxl.load_workbook(self.file_path, data_only=True)
        self.worksheet = self.workbook.active

    def get_all_teachers(self) -> List[str]:
        """获取所有教师姓名"""
        teachers = set()
        for row in self.worksheet.iter_rows(values_only=True):
            for cell in row:
                if cell and isinstance(cell, str):
                    # 匹配格式: Фамилия И.О.
                    matches = re.findall(r'[А-Яа-яЁё]+\s+[А-Я]\.[А-Й]\.', cell)
                    teachers.update(matches)
        # 排除一些误匹配
        teachers = {t for t in teachers if 'наследие' not in t and len(t.split()) <= 3}
        return sorted(teachers)

    def parse_cell(self, cell_value: str) -> Optional[dict]:
        """解析单元格内容，提取课程信息"""
        if not cell_value or not isinstance(cell_value, str):
            return None

        # 提取教师姓名
        teacher_match = re.search(r'([А-Яа-яЁё]+\s+[А-Я]\.[А-Й]\.)', cell_value)
        teacher = teacher_match.group(1) if teacher_match else ''

        # 提取教室
        room_match = re.search(r'(\d+-\d+[а-я]?)', cell_value)
        room = room_match.group(1) if room_match else ''
        if not room:
            room_match = re.search(r'([А-Я]+-\d+[а-я]?)', cell_value)
            room = room_match.group(1) if room_match else ''

        # 提取活动类型
        activity_type = ''
        if '(лк)' in cell_value or '(лекция)' in cell_value.lower():
            activity_type = 'лекция'
        elif '(пз)' in cell_value or '(практическое)' in cell_value.lower():
            activity_type = 'практическое'
        elif '(лб)' in cell_value or '(лабораторная)' in cell_value.lower():
            activity_type = 'лабораторная'

        # 提取科目名称（在活动类型之前的部分）
        subject = cell_value
        if activity_type:
            subject = subject.split('(')[0].strip()

        return {
            'teacher': teacher,
            'room': room,
            'subject': subject,
            'activity_type': activity_type
        }

    def get_teacher_schedule(self, teacher_name: str) -> List[ScheduleEntry]:
        """获取指定教师的课程表"""
        schedule = []
        current_day = None
        current_time = None
        current_groups = []

        # 首先获取所有组别
        for col in self.worksheet.iter_cols(min_row=3, max_row=3, values_only=True):
            if col[0] and isinstance(col[0], str):
                if 'ПМ-' in col[0] or 'ИиВТ-' in col[0]:
                    current_groups.append(col[0].strip())

        for row_idx, row in enumerate(self.worksheet.iter_rows(values_only=True), 1):
            # 检查是否是星期行
            for cell in row[:3]:
                if cell and isinstance(cell, str):
                    # 去除空格后检查是否是星期
                    clean_cell = cell.replace(' ', '')
                    for day_key, day_value in self.DAYS_MAP.items():
                        if day_key in clean_cell:
                            current_day = day_value
                            break

            # 检查时间列（通常是第二列）
            if len(row) > 1 and row[1]:
                time_cell = str(row[1])
                if re.match(r'\d+-\d+', time_cell):
                    current_time = time_cell

            # 搜索教师的所有单元格
            if current_day and current_time:
                for col_idx, cell in enumerate(row[2:], 2):
                    if cell and isinstance(cell, str) and teacher_name in cell:
                        parsed = self.parse_cell(cell)
                        if parsed:
                            # 获取组别（从表头或列索引推断）
                            group = current_groups[0] if current_groups else ''

                            entry = ScheduleEntry(
                                day=current_day,
                                time=current_time,
                                subject=parsed['subject'],
                                teacher=teacher_name,
                                room=parsed['room'],
                                group=group,
                                activity_type=parsed['activity_type']
                            )
                            schedule.append(entry)

        return schedule

    def get_all_schedules(self) -> Dict[str, List[ScheduleEntry]]:
        """获取所有教师的课程表"""
        teachers = self.get_all_teachers()
        schedules = {}

        for teacher in teachers:
            schedule = self.get_teacher_schedule(teacher)
            if schedule:
                schedules[teacher] = schedule

        return schedules


if __name__ == '__main__':
    reader = ExcelReader('ITsTiM_Raspisanie_2_polugodie_25-26_mag__pechat.xlsx')
    reader.load()

    teachers = reader.get_all_teachers()
    print(f"找到 {len(teachers)} 位教师:")
    for t in teachers:
        print(f"  - {t}")
